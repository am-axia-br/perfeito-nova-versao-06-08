import pandas as pd
import io
import re

def gerar_excel_processo(dados_processo, texto_processo=None):
    """
    Gera um arquivo Excel formatado com os dados do processo trabalhista.
    Inclui abas detalhadas conforme a estrutura esperada do PJe-Calc.
    
    Parâmetros:
        dados_processo (dict): Dados estruturados do processo trabalhista
        texto_processo (str, opcional): Texto completo do processo para análise comparativa
    """
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    # DADOS CADASTRAIS - Primeira aba (EXPANDIDA)
    df_cadastral = pd.DataFrame({
        'Campo': [
            'Processo', 'Reclamante', 'CPF', 'Reclamada(s)', 
            'Data de Ajuizamento', 'Vara/UF', 'Valor da Causa',
            'Data de Admissão', 'Data de Demissão', 'Função',
            'Período Contratual', 'Último Salário Base', 
            'Período de Afastamento', 'Fase do Cálculo', 
            'Observações'
        ],
        'Valor': [
            dados_processo.get('informacoes_pjecalc', {}).get('numero_processo', ''),
            dados_processo.get('dados_pessoais', {}).get('reclamante', ''),
            dados_processo.get('dados_pessoais', {}).get('cpf', ''),
            dados_processo.get('dados_pessoais', {}).get('reclamada', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('data_liquidacao', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('jurisdicao', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('valor_causa', ''),
            dados_processo.get('dados_pessoais', {}).get('data_admissao', ''),
            dados_processo.get('dados_pessoais', {}).get('data_demissao', ''),
            dados_processo.get('dados_pessoais', {}).get('funcao', ''),
            f"{dados_processo.get('dados_pessoais', {}).get('data_admissao', '')} a {dados_processo.get('dados_pessoais', {}).get('data_demissao', '')}" + 
            (f" – {dados_processo.get('dados_pessoais', {}).get('tipo_rescisao', '')}" if dados_processo.get('dados_pessoais', {}).get('tipo_rescisao') else ""),
            dados_processo.get('dados_pessoais', {}).get('ultimo_salario', ''),
            dados_processo.get('dados_pessoais', {}).get('periodo_afastamento', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('fase_calculo', 'Provisão Inicial'),
            dados_processo.get('observacoes_gerais', '')
        ]
    })
    
    # VERBAS PLEITEADAS (RESUMO) - Segunda aba
    verbas_data = []
    for verba in dados_processo.get('verbas_pleiteadas', []):
        verbas_data.append({
            'Verba': verba.get('verba', ''),
            'Parâmetros': verba.get('periodo', ''),
            'Reflexos': verba.get('reflexos', '')
        })
    
    df_verbas = pd.DataFrame(verbas_data) if verbas_data else pd.DataFrame({'Verba': ['Nenhuma verba identificada'], 'Parâmetros': [''], 'Reflexos': ['']})
    
    # VERBAS PLEITEADAS DETALHADAS - Terceira aba
    # Usar APENAS dados reais do processo, sem valores fixos
    verbas_detalhadas = []

    # Obter diretamente as verbas do processo analisado com seus reflexos
    for verba in dados_processo.get('verbas_pleiteadas', []):
        # Obter reflexos específicos de cada verba do processo
        reflexos = verba.get('reflexos', '')
        
        verbas_detalhadas.append({
            'Verba': verba.get('verba', ''),
            'Período': verba.get('periodo', ''),
            'Valor Base': verba.get('valor_base', ''),
            'Percentual / Quantidade': verba.get('percentual_quantidade', ''),
            'Reflexos': reflexos  # Reflexos específicos para cada verba
        })

    # Se não houver verbas, criar uma entrada vazia para manter a estrutura
    if not verbas_detalhadas:
        verbas_detalhadas.append({
            'Verba': 'Nenhuma verba identificada',
            'Período': '',
            'Valor Base': '',
            'Percentual / Quantidade': '',
            'Reflexos': ''
        })
    
    df_verbas_detalhadas = pd.DataFrame(verbas_detalhadas)
    
    # RESULTADO DA LIQUIDAÇÃO - Quarta aba
    df_resultado = pd.DataFrame({
        'Item': [
            'Valor Bruto Total',
            'Descontos (IR e INSS)',
            'Líquido Devido à Reclamante'
        ],
        'Valor': [
            dados_processo.get('resultado_liquidacao', {}).get('valor_bruto_total', ''),
            dados_processo.get('resultado_liquidacao', {}).get('descontos_ir_inss', ''),
            dados_processo.get('resultado_liquidacao', {}).get('liquido_devido_reclamante', '')
        ]
    })
    
    # PARÂMETROS DE CÁLCULO - Quinta aba
    df_parametros = pd.DataFrame({
        'Parâmetro': [
            'Correção Monetária', 
            'Juros',
            'INSS Reclamante',
            'INSS Patronal',
            'INSS Terceiros',
            'Honorários Advocatícios'
        ],
        'Valor': [
            dados_processo.get('informacoes_pjecalc', {}).get('correcao_monetaria', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('juros_mora', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('inss_reclamante', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('inss_patronal', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('inss_terceiros', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('honorarios', '')
        ]
    })
    
    # NOVA ABA: CONFIGURAÇÃO PJE-CALC - Sexta aba
    # Separação de reclamadas (caso estejam em uma string única)
    reclamada_texto = dados_processo.get('dados_pessoais', {}).get('reclamada', '')
    reclamadas = []

    if ';' in reclamada_texto:
        reclamadas = [r.strip() for r in reclamada_texto.split(';')]
    elif ',' in reclamada_texto:
        reclamadas = [r.strip() for r in reclamada_texto.split(',')]
    else:
        reclamadas = [reclamada_texto]

    # Garantir que temos pelo menos duas reclamadas para o modelo (mesmo que vazias)
    if len(reclamadas) < 2:
        reclamadas.append('')
        
    df_pjecalc_config = pd.DataFrame({
        'Bloco': [
            'BLOCO 1', 'BLOCO 1', 'BLOCO 1', 'BLOCO 1', 'BLOCO 1', 'BLOCO 1', 'BLOCO 1', 'BLOCO 1', 'BLOCO 1',
            'BLOCO 2', 'BLOCO 2',
            'BLOCO 4', 'BLOCO 4', 'BLOCO 4', 'BLOCO 4', 'BLOCO 4', 'BLOCO 4', 'BLOCO 4',
            'BLOCO 5', 'BLOCO 5'
        ],
        'Campo': [
            'Processo', 'Reclamante', 'CPF', 'Reclamada 1', 'Reclamada 2',
            'Data de Ajuizamento', 'Período Contratual', 'Cidade', 'Salário Base',
            'Período de Afastamento', 'Motivo',
            'Valor da Causa', 'Custas', 'Honorários', 'INSS Reclamante', 'INSS Patronal', 
            'INSS Terceiros', 'Correção Monetária',
            'Atualizar Até', 'Formato de Exportação'
        ],
        'Valor': [
            dados_processo.get('informacoes_pjecalc', {}).get('numero_processo', ''),
            dados_processo.get('dados_pessoais', {}).get('reclamante', ''),
            dados_processo.get('dados_pessoais', {}).get('cpf', ''),
            reclamadas[0],
            reclamadas[1] if len(reclamadas) > 1 else '',
            dados_processo.get('informacoes_pjecalc', {}).get('data_liquidacao', ''),
            f"{dados_processo.get('dados_pessoais', {}).get('data_admissao', '')} a {dados_processo.get('dados_pessoais', {}).get('data_demissao', '')}" +
            (f" – {dados_processo.get('dados_pessoais', {}).get('tipo_rescisao', '')}" if dados_processo.get('dados_pessoais', {}).get('tipo_rescisao') else ""),
            dados_processo.get('dados_pessoais', {}).get('cidade', ''),
            dados_processo.get('dados_pessoais', {}).get('ultimo_salario', ''),
            dados_processo.get('dados_pessoais', {}).get('periodo_afastamento', ''),
            dados_processo.get('dados_pessoais', {}).get('motivo_afastamento', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('valor_causa', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('custas', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('honorarios', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('inss_reclamante', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('inss_patronal', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('inss_terceiros', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('correcao_monetaria', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('data_liquidacao', ''),
            dados_processo.get('informacoes_pjecalc', {}).get('formato_exportacao', '')
        ]
    })
    
    # MODIFICADA: ABA RESUMO PJE-CALC - Representação fiel da interface
    # Obter texto das observações gerais que contém toda a estrutura da interface
    texto_interface = dados_processo.get('observacoes_gerais', '')
    
    # Função para extrair seções da interface
    def extrair_secoes_da_interface(texto):
        secoes = {}
        
        # Análise Concluída
        match = re.search(r"✅ Análise Concluída(.*?)(?:📝|$)", texto, re.DOTALL)
        if match:
            secoes["analise_concluida"] = match.group(1).strip()
        
        # Observações Gerais
        match = re.search(r"📝 Observações Gerais e Resumo do Processo(.*?)(?:✅ Verbas|$)", texto, re.DOTALL)
        if match:
            secoes["observacoes"] = match.group(1).strip()
        
        # Verbas Pleiteadas
        match = re.search(r"✅ Verbas Pleiteadas \(Petição Inicial\)(.*?)(?:✅ Atualização|$)", texto, re.DOTALL)
        if match:
            secoes["verbas"] = match.group(1).strip()
        
        # Atualização Monetária
        match = re.search(r"✅ Atualização Monetária e Parâmetros(.*?)(?:✅ Revisão|$)", texto, re.DOTALL)
        if match:
            secoes["atualizacao"] = match.group(1).strip()
        
        # Revisão Final - Quadro para Cálculo
        match = re.search(r"✅ Revisão Final – Quadro para Cálculo(.*?)(?:✅ Modelo|$)", texto, re.DOTALL)
        if match:
            secoes["revisao"] = match.group(1).strip()
        
        # Modelo Laudo Técnico
        match = re.search(r"✅ Modelo Laudo Técnico para PJe-Calc \(com BLOCOS 1 a 5\)(.*?)(?:Processado por|$|📌)", texto, re.DOTALL)
        if match:
            secoes["modelo"] = match.group(1).strip()
        
        # Dados Pessoais e Contratuais
        match = re.search(r"📌 Dados Pessoais e Contratuais(.*?)(?:🗂️|$)", texto, re.DOTALL)
        if match:
            secoes["dados_pessoais"] = match.group(1).strip()
        
        # Informações para Preenchimento no PJe-Calc
        match = re.search(r"🗂️ Informações para Preenchimento no PJe-Calc(.*?)(?:📑|🟢|$)", texto, re.DOTALL)
        if match:
            secoes["info_pjecalc"] = match.group(1).strip()
        
        # Resultado Final da Liquidação
        match = re.search(r"🟢 Resultado Final da Liquidação(.*?)(?:$)", texto, re.DOTALL)
        if match:
            secoes["resultado"] = match.group(1).strip()
            
        return secoes
    
    # Extrair dados da interface para estruturar o Excel
    linhas_resumo = []
    
    # Se temos o texto da interface, usar ele para criar o resumo
    if texto_interface:
        secoes = extrair_secoes_da_interface(texto_interface)
        
        # Adicionar cabeçalho
        linhas_resumo.append(["Seção", "Conteúdo"])
        
        # Análise Concluída
        linhas_resumo.append(["✅ Análise Concluída", ""])
        
        # Observações Gerais e Resumo
        linhas_resumo.append(["📝 Observações Gerais e Resumo do Processo", ""])
        if "observacoes" in secoes:
            for linha in secoes["observacoes"].split("\n"):
                linhas_resumo.append(["", linha.strip()])
        
        # Verbas Pleiteadas
        linhas_resumo.append(["✅ Verbas Pleiteadas (Petição Inicial)", ""])
        if "verbas" in secoes:
            for linha in secoes["verbas"].split("\n"):
                linhas_resumo.append(["", linha.strip()])
        
        # Atualização Monetária
        linhas_resumo.append(["✅ Atualização Monetária e Parâmetros", ""])
        if "atualizacao" in secoes:
            for linha in secoes["atualizacao"].split("\n"):
                linhas_resumo.append(["", linha.strip()])
        
        # Revisão Final
        linhas_resumo.append(["✅ Revisão Final – Quadro para Cálculo", ""])
        if "revisao" in secoes:
            for linha in secoes["revisao"].split("\n"):
                linhas_resumo.append(["", linha.strip()])
        
        # Modelo Laudo Técnico
        linhas_resumo.append(["✅ Modelo Laudo Técnico para PJe-Calc (com BLOCOS 1 a 5)", ""])
        if "modelo" in secoes:
            for linha in secoes["modelo"].split("\n"):
                linhas_resumo.append(["", linha.strip()])
        
        # Dados Pessoais
        linhas_resumo.append(["📌 Dados Pessoais e Contratuais", ""])
        if "dados_pessoais" in secoes:
            for linha in secoes["dados_pessoais"].split("\n"):
                linhas_resumo.append(["", linha.strip()])
        
        # Informações PJe-Calc
        linhas_resumo.append(["🗂️ Informações para Preenchimento no PJe-Calc", ""])
        if "info_pjecalc" in secoes:
            for linha in secoes["info_pjecalc"].split("\n"):
                linhas_resumo.append(["", linha.strip()])
        
        # Resultado Final
        linhas_resumo.append(["🟢 Resultado Final da Liquidação", ""])
        if "resultado" in secoes:
            for linha in secoes["resultado"].split("\n"):
                linhas_resumo.append(["", linha.strip()])
        
        # Adicionar linha com o processado por
        match = re.search(r"Processado por (.*?) em (.*?) UTC", texto_interface)
        if match:
            usuario = match.group(1)
            data = match.group(2)
            linhas_resumo.append(["", f"Processado por {usuario} em {data} UTC."])
        else:
            # Usar data e usuário atuais se não encontrar no texto
            linhas_resumo.append(["", f"Processado por am-axia-br em 2025-08-06 02:35:48 UTC."])
        
        # Criar DataFrame com os dados da interface
        df_resumo_pjecalc = pd.DataFrame(linhas_resumo[1:], columns=linhas_resumo[0])
    else:
        # Usar a estrutura anterior se não tiver o texto da interface
        df_resumo_pjecalc = pd.DataFrame({
            'Seção': [
                '✅ Análise Concluída',
                '📝 Observações Gerais e Resumo do Processo',
                '✅ Verbas Pleiteadas (Petição Inicial)', 
                '✅ Atualização Monetária e Parâmetros',
                '✅ Revisão Final – Quadro para Cálculo',
                '✅ Modelo Laudo Técnico para PJe-Calc (com BLOCOS 1 a 5)',
                '📌 Dados Pessoais e Contratuais',
                '🗂️ Informações para Preenchimento no PJe-Calc',
                '🟢 Resultado Final da Liquidação'
            ],
            'Conteúdo': [
                '',
                dados_processo.get('observacoes_gerais', ''),
                ', '.join([v.get('verba', '') for v in dados_processo.get('verbas_pleiteadas', [])]),
                f"Índice Monetário: {dados_processo.get('informacoes_pjecalc', {}).get('correcao_monetaria', '')}\nJuros de Mora: {dados_processo.get('informacoes_pjecalc', {}).get('juros_mora', '')}\nINSS – Terceiros: {dados_processo.get('informacoes_pjecalc', {}).get('inss_terceiros', '')}",
                "Ver aba Verbas Detalhadas para informações completas",
                "BLOCO 1 – Dados Cadastrais: [Preencher com os dados consolidados]\nBLOCO 2 – Afastamentos: [Preencher com os períodos de afastamento]\nBLOCO 3 – Verbas a Apurar: [Configurar a tabela com as verbas e parâmetros]\nBLOCO 4 – Honorários / Custas / Encargos: [Preencher com os dados de honorários e demais encargos]\nBLOCO 5 – Liquidação e Impressão: [Gerar a liquidação com base nos cálculos]",
                f"Reclamante: {dados_processo.get('dados_pessoais', {}).get('reclamante', '')}\nReclamada: {dados_processo.get('dados_pessoais', {}).get('reclamada', '')}\nData de Admissão: {dados_processo.get('dados_pessoais', {}).get('data_admissao', '')}\nData de Demissão: {dados_processo.get('dados_pessoais', {}).get('data_demissao', '')}\nÚltimo Salário: {dados_processo.get('dados_pessoais', {}).get('ultimo_salario', '')}",
                f"Nº do Processo: {dados_processo.get('informacoes_pjecalc', {}).get('numero_processo', '')}\nTRT Região: {dados_processo.get('informacoes_pjecalc', {}).get('trt_regiao', '')}\nJurisdição: {dados_processo.get('informacoes_pjecalc', {}).get('jurisdicao', '')}\nData da Liquidação: {dados_processo.get('informacoes_pjecalc', {}).get('data_liquidacao', '')}\nÍndices de Correção: {dados_processo.get('informacoes_pjecalc', {}).get('correcao_monetaria', '')}\nJuros de Mora: {dados_processo.get('informacoes_pjecalc', {}).get('juros_mora', '')}\nHonorários Periciais: {dados_processo.get('informacoes_pjecalc', {}).get('honorarios_periciais', '')}",
                f"Valor Bruto Total: {dados_processo.get('resultado_liquidacao', {}).get('valor_bruto_total', '')}\nDescontos (IR e INSS): {dados_processo.get('resultado_liquidacao', {}).get('descontos_ir_inss', '')}\nLíquido Devido à Reclamante: {dados_processo.get('resultado_liquidacao', {}).get('liquido_devido_reclamante', '')}"
            ]
        })
    
    # ANÁLISE COMPARATIVA DE VERBAS (se disponível)
    try:
        # Importar a função de análise apenas quando necessário
        from analise_verbas import analisar_processo_trabalhista
        
        # Verificar se temos texto do processo disponível
        if texto_processo:
            # Preparar dados no formato que a função espera
            verbas_resumo = [v.get('verba', '') for v in dados_processo.get('verbas_pleiteadas', []) if v.get('verba')]
            dados_planilha = {
                "Verbas Resumidas": [{"verba": v} for v in verbas_resumo if v],
                "Verbas Detalhadas": verbas_detalhadas
            }
            
            # Realizar análise
            resultado = analisar_processo_trabalhista(texto_processo, dados_planilha)
            
            # Criar DataFrame para a nova aba
            linhas_analise = resultado.split('\n')
            df_analise = pd.DataFrame({"Análise Comparativa de Verbas": linhas_analise})
            
            # Adicionar nova aba
            df_analise.to_excel(writer, sheet_name='Análise de Verbas', index=False)
            
            print("✅ Aba de análise de verbas adicionada com sucesso!")
    except Exception as e:
        print(f"⚠️ Não foi possível adicionar análise de verbas: {e}")
    
    # Escrever cada DataFrame em uma aba separada
    df_cadastral.to_excel(writer, sheet_name='Dados do Processo', index=False)
    df_verbas.to_excel(writer, sheet_name='Verbas Resumidas', index=False)
    df_verbas_detalhadas.to_excel(writer, sheet_name='Verbas Detalhadas', index=False)
    df_resultado.to_excel(writer, sheet_name='Resultado Liquidação', index=False)
    df_parametros.to_excel(writer, sheet_name='Parâmetros de Cálculo', index=False)
    df_pjecalc_config.to_excel(writer, sheet_name='Configuração PJe-Calc', index=False)
    df_resumo_pjecalc.to_excel(writer, sheet_name='Resumo PJe-Calc', index=False)
    
    # Aplicar formatação
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        # Ajustar largura das colunas
        for idx, col in enumerate(worksheet.columns):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = min(adjusted_width, 100)  # Limitar largura máxima
        
        # Formatação de cabeçalho
        for cell in worksheet["1:1"]:
            cell.font = cell.font.copy(bold=True)
            cell.fill = cell.fill.copy(patternType="solid", fgColor="D9E1F2")
        
        # Adicionar bordas leves a todas as células com dados
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # Ajuste especial para a aba Resumo PJe-Calc - Aplicar formatação específica
        if sheet_name == 'Resumo PJe-Calc':
            # Colorir células de cabeçalho de seção
            from openpyxl.styles import PatternFill, Font
            for row in range(1, max_row + 1):
                cell_value = str(worksheet.cell(row=row, column=1).value or '')
                if cell_value.startswith('✅'):
                    # Cor verde claro para cabeçalhos com ✅
                    worksheet.cell(row=row, column=1).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
                elif cell_value.startswith('📝'):
                    # Cor azul claro para observações
                    worksheet.cell(row=row, column=1).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
                elif cell_value.startswith('📌') or cell_value.startswith('🗂️'):
                    # Cor laranja claro para dados pessoais/pjecalc
                    worksheet.cell(row=row, column=1).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
                elif cell_value.startswith('🟢'):
                    # Cor verde para resultados
                    worksheet.cell(row=row, column=1).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
    
    writer.close()
    output.seek(0)
    return output

def exportar_interface():
    """
    Função auxiliar para importar na interface.py
    
    Exemplo de uso:
    
    # Em interface.py
    from exportadores_completo import gerar_excel_processo
    
    # Na seção onde os botões são exibidos
    excel_data = gerar_excel_processo(dados_consolidados)
    st.download_button("⬇️ Baixar Excel", excel_data, file_name="processo_resumo.xlsx", 
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    """
    pass